import os
import re
import warnings
from datetime import datetime

import dateutil
import pandas as pd
from openpyxl import load_workbook


QUIET = False

def valid_times(cells):
    """Get the valid hours for the excel spreadsheet

    The columns are not always the same on different files. This is why finding the headers for the times was a good solution to find the correct columns were data should be extracted.
    """

    hours = []
    for row in cells.rows:
        for cell in row:
            if cell.row == 12: # 12 is hardcoded because it is the row where the time headers are found
                if cell.value is not None:
                    # print(cell.value)
                    hours.append(cell)

    return hours


def throughput(cells, times):
    """Get the throughput for all columns"""

    ride_throughput = {}
    for time in times:
        value_coordinate = str(time.column) + '20' # 20 is hardcoded to the row number that corresponds to the Ride Throughput
        value = cells[value_coordinate].value
        if value != 'Ride Throughput':
            ride_throughput[time.value] = value

    return ride_throughput


def get_files(directory, inc_ext=['xlsx']):
    """Gets file names of file types from a directory"""

    file_names = [
            filename \
            for filename in os.listdir(os.path.realpath(directory)) \
            if any(filename.endswith(ext) for ext in inc_ext)
            ]

    return file_names


def sweep_cells(cells):
    times = valid_times(cells)
    pkt = throughput(cells, times)

    return pkt


def sweep_sheets(workbook, date):
    """Looks through the sheets of a given document"""

    sheet_names = workbook.sheetnames

    rows = []

    for sheet_name in sheet_names:
        cells = workbook[sheet_name]
        ride_name = cells['A6'].value # A6 is hardcoded to the ride_name on the sheet
        pkt = sweep_cells(cells)

        row = {'ride_name': ride_name, 'date': date }

        # TODO: replace this with more pythonic solution
        # TODO: figure out why row.update({}) was rendering None
        for time, value in pkt.items():
            row[time]=value

        rows.append(row)

    return rows



def sweep_documents(directory):
    """Looks through the documents of a given document"""
    global QUIET
    file_names = get_files(directory)

    data = []
    for filename in file_names:
        if not QUIET:
            print('Processing: ' + filename)
        wb = load_workbook(filename = os.path.join(directory, filename))

        # extract date from filename
        raw_date = re.sub('[DW_.xlsx]', '', filename)
        date = dateutil.parser.parse(raw_date)

        data.extend(sweep_sheets(wb, date))

    return data


def save(df, filename):
    """Write DataFrame to a file"""

    if filename.endswith('xlsx'):
        writer = pd.ExcelWriter(filename, engine='xlsxwriter')
        df.to_excel(writer)
        writer.save()
    elif filename.endswith('csv'):
        df.to_csv(filename, sep=',')
    else:
        # TODO: Have the warning print the filetype
        ext = os.path.splitext(filename)[1]
        warnings.warn('Cannot save filetype ' + ext )


def run():
    """Sweeps through documents to extract information regarding the throughput
    of rides

    Returns a DataFrame
    """
    global QUIET
    QUIET = True
    return main()


def fix_titles(df, original, replacement):
    """Replaces source titles with ones that make more standardized times"""

    new_columns = {old:new for old, new in zip(original, replacement)}

    return df.rename(columns=new_columns)


def flatten_time(df, times):
    """Organizes the dataframe to eliminate having time in separate columns

    Moves all the column hours to the date columns, creating separate entries for each one.
    """

    rows = []
    for index, row in df.iterrows():
        for time in times:
            new_time = datetime.strptime(time, '%I:%M %p').time()
            new_date = datetime.combine(datetime.date(row['date']), new_time)
            rows.append({'ride_name':row['ride_name'],'date':new_date,'throughput':row[time]})

    flat = pd.DataFrame(rows, columns=['ride_name','date','throughput'])
    return flat


def main():
    """Sweeps through documents to extract information regarding the throughput of rides"""

    wk_dir = os.path.dirname(os.path.realpath('__file__'))
    directory = os.path.join(wk_dir, 'resources/attraction-operational-readiness-reports')

    # TODO: Add working debug mode
    # directory = os.path.join(wk_dir, 'test_files')

    # Most of the work happens here:
    data = sweep_documents(directory)

    source_times = [
                  '9a-10a',
                  '10a-11a',
                  '11a-12p',
                  '12p-1p',
                  '1p-2p',
                  '2p-3p',
                  '3p-4p',
                  '4p-5p',
                  '5p-6p',
                  '6p-7p',
                  '7p-8p',
                  '8p-9p',
                  '9p-10p',
                  '10p-11p'
                  ]

    times = [
             '10:00 am',
             '11:00 am',
             '12:00 pm',
             '01:00 pm',
             '02:00 pm',
             '03:00 pm',
             '04:00 pm',
             '05:00 pm',
             '06:00 pm',
             '07:00 pm',
             '08:00 pm',
             '09:00 pm',
             '10:00 pm',
             '11:00 pm'
             ]

    source_titles = ['ride_name', 'date']
    source_titles.extend(source_times)

    df = pd.DataFrame(data, columns=source_titles)

    old_format = df[source_titles].sort_values(by=['ride_name', 'date'])
    human_readable = fix_titles(old_format, source_times, times)
    logic = flatten_time(human_readable, times)

    assert len(logic.index) == len(human_readable.index)*14, 'Lengths do not match'

    save(logic, 'output/dataframe.csv')
    save(human_readable, 'output/human-readable.xlsx')
    save(old_format, 'output/old_format.xlsx')

    return logic


if __name__ == '__main__':
    # TODO: Add argparser
    main()

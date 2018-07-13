import os

import pandas as pd
from openpyxl import load_workbook


RIDES = {}
ROWS = []



#TODO: Remove, ended up not using these classes

class day():
    def __init__(self, date, throughput):
        self.date = date
        self.throughput = throughput
        # {'9a-10a':234, ...}
    def __repr__(self): 
        return str(self.date) + ' ' + str(self.throughput)


class ride():
    def __init__(self, name):
        self.name = name
        self.days = []
        # {'DW_102117.xlsx':[]}


def make_new():
    from openpyxl import Workbook
    from openpyxl.compat import range
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    dest_filename = 'empty_book.xlsx'

    ws1 = wb.active
    ws1.title = "range names"

    for row in range(1, 40):
        ws1.append(range(600))

    ws2 = wb.create_sheet(title="Pi")

    ws2['F5'] = 3.14

    ws3 = wb.create_sheet(title="Data")
    for row in range(10, 20):
        for col in range(27, 54):
            _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
    print(ws3['AA10'].value)

    wb.save(filename = dest_filename)


def valid_times(cells):
    """Get the valid hours for an excel spreadsheet"""
    hours = []
    for row in cells.rows:
        for cell in row:
            if cell.row == 12:
                if cell.value is not None:
                    # print(cell.value)
                    hours.append(cell)
    return hours


def throughput(cells, times):
    """Get the throughput for all columns"""
    ride_throughput = {}
    for time in times:
        value_coordinate = str(time.column) + '20'
        value = cells[value_coordinate].value
        if value != 'Ride Throughput':
            ride_throughput[time.value] = value
            # print('ride_throughput ,', time.value, ',',  value)

    return ride_throughput


def first_throughput(cells):
    """Proof of concept for throughput"""
    tp = cells['D20'].value

    if tp is None:
        tp = cells['E20'].value

    return tp


def print_points(sheet_name, ride_name, throughput):
    """Pretty print"""
    print()
    print()
    print('sheet_name ,', sheet_name)
    print('ride_name ,', ride_name)

    for elem in throughput:
        print('throughput ,' , elem[0], ',', elem[1])


def get_files(directory, inc_ext=['xlsx']):
    """Gets file names of file types from a directory"""
    file_names = [
            filename \
            for filename in os.listdir(os.path.realpath(directory)) \
            if any(filename.endswith(ext) for ext in inc_ext)
            ]
    return file_names


def sweep_cells(cells):
    times = valid_times(cells)
    tp = throughput(cells, times)
    return tp


    def to_excel(self):
        return {self.name:self.days}


def sweep_sheets(workbook, date):
    # sheet_names = ['Sheet' + str(i) for i in range(1,32)]
    sheet_names = workbook.sheetnames

    for sheet_name in sheet_names:
        cells = workbook[sheet_name]
        ride_name = cells['A6'].value
        tp = sweep_cells(cells)

        row = {'ride_name': ride_name, 'date': date }
        for k, v in tp.items():
            row[k]=v
        ROWS.append(row)

        #TODO: Remove
        # first time run creates objects
        # if ride_name not in RIDES:
        #     RIDES[ride_name] = ride(ride_name)
        # print(ride_name)
        # print(date)
        # print(tp)
        # RIDES[ride_name].days.append(day(date, tp))


def sweep_documents(directory):
    file_names = get_files(directory)
    for filename in file_names:
        print()
        print('--------------------------------------------')
        print('Loading ' + filename)
        print('--------------------------------------------')
        wb = load_workbook(filename = os.path.join(directory, filename))
        import re
        import dateutil

        raw_date = re.sub('[DW_.xlsx]', '', filename)
        date = dateutil.parser.parse(raw_date)
        sweep_sheets(wb, date)


def save(df, filename):
    # df.to_csv(filename)
    df.to_excel(filename)



def main():

    sweep_documents('attraction-operational-readiness-reports')
    # sweep_documents('test')

    df = pd.DataFrame(ROWS)

    TITLES = ['ride_name',
              'date',
              '9a-10a',
              '10a-11a',
              '11a-12p',
              '12p-1p',
              '1p-2p',
              '2p-3p',
              '3p-4p',
              '4p-5p',
              '5p-6p',
              '6p-7p',
              '7p-8p',
              '8p-9p',
              '9p-10p',
              '10p-11p'
              ]

    df = df[TITLES]
    sorted_df = df.sort_values(by=['ride_name','date'])
    save(sorted_df, 'testing.xlsx')



if __name__ == '__main__':
    main()
